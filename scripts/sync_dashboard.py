#!/usr/bin/env python3
"""
sync_dashboard.py
-----------------
Low-code helper to keep DASHBOARD.xlsx in sync with slide work.

Typical usage:
  python scripts/sync_dashboard.py --deck GRADE --slides S06-S15 S03 \
    --status READY --p0 No --pdf_ok Yes --evidence_ok Yes --visual 8 \
    --notes "Batch 06–15: contraste/hierarquia + cards GRADE em layout claro" \
    --title-from-html

Notes:
- Uses openpyxl only (no external HTML parser).
- Will update the "Slides" sheet (expects the standard header row).
"""

from __future__ import annotations

import argparse
import datetime as _dt
import html as _html
import os
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl


def parse_slide_tokens(tokens: List[str]) -> List[str]:
    """
    Accepts tokens like:
      S03
      S06-S15
      s09-s13
    Returns a list of normalized slide IDs: ["S03","S06",...]
    """
    out: List[str] = []
    for tok in tokens:
        tok = tok.strip()
        if not tok:
            continue
        tok = tok.upper()
        if "-" in tok:
            a, b = tok.split("-", 1)
            a = a.strip().upper()
            b = b.strip().upper()
            ma = re.match(r"^S(\d+)$", a)
            mb = re.match(r"^S(\d+)$", b)
            if not (ma and mb):
                raise ValueError(f"Invalid slide range: {tok}")
            ia = int(ma.group(1))
            ib = int(mb.group(1))
            if ia > ib:
                ia, ib = ib, ia
            for i in range(ia, ib + 1):
                out.append(f"S{i:02d}")
        else:
            if not re.match(r"^S\d+$", tok):
                raise ValueError(f"Invalid slide id: {tok}")
            # normalize to 2 digits when possible
            m = re.match(r"^S(\d+)$", tok)
            if m:
                out.append(f"S{int(m.group(1)):02d}")
            else:
                out.append(tok)
    # de-duplicate while preserving order
    seen = set()
    out2 = []
    for s in out:
        if s not in seen:
            out2.append(s)
            seen.add(s)
    return out2


def extract_title_from_html(path: Path) -> str:
    if not path.exists():
        return ""
    txt = path.read_text(encoding="utf-8", errors="ignore")

    # Prefer h1, then h2
    for tag in ("h1", "h2"):
        m = re.search(rf"<{tag}[^>]*>(.*?)</{tag}>", txt, flags=re.IGNORECASE | re.DOTALL)
        if not m:
            continue
        inner = m.group(1)
        # strip nested tags
        inner = re.sub(r"<[^>]+>", " ", inner)
        inner = _html.unescape(inner)
        inner = " ".join(inner.split())
        return inner
    return ""


def build_header_map(ws) -> Dict[str, int]:
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if isinstance(v, str) and v.strip():
            headers[v.strip()] = c
    return headers


def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="DASHBOARD.xlsx", help="Path to DASHBOARD.xlsx (default: ./DASHBOARD.xlsx)")
    ap.add_argument("--deck", default="GRADE", help="Deck_ID to update (default: GRADE)")
    ap.add_argument("--slides", nargs="+", required=True, help="Slide IDs or ranges, e.g. S06-S15 S03")
    ap.add_argument("--status", default=None, help="Set Status (e.g., READY, DRAFT)")
    ap.add_argument("--p0", default=None, help="Set P0 (Yes/No)")
    ap.add_argument("--pdf_ok", default=None, help="Set PDF_OK (Yes/No)")
    ap.add_argument("--evidence_ok", default=None, help="Set Evidence_OK (Yes/No)")
    ap.add_argument("--visual", type=int, default=None, help="Set Visual_0_10 (0-10)")
    ap.add_argument("--reviewed", default=None, help="Set Last_Reviewed (YYYY-MM-DD). Default: today.")
    ap.add_argument("--notes", default=None, help="Append to Notes column")
    ap.add_argument("--title-from-html", action="store_true", help="Update Title from slide HTML (h1/h2).")
    args = ap.parse_args(argv)

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"ERROR: XLSX not found: {xlsx_path}", file=sys.stderr)
        return 2

    slide_ids = parse_slide_tokens(args.slides)
    reviewed = args.reviewed or _dt.date.today().isoformat()

    wb = openpyxl.load_workbook(xlsx_path)
    if "Slides" not in wb.sheetnames:
        print("ERROR: Sheet 'Slides' not found in workbook.", file=sys.stderr)
        return 3

    ws = wb["Slides"]
    hmap = build_header_map(ws)

    required = ["Deck_ID", "Slide_ID"]
    for k in required:
        if k not in hmap:
            print(f"ERROR: Missing required column header: {k}", file=sys.stderr)
            return 4

    updated = 0
    missing: List[str] = []

    for sid in slide_ids:
        row_idx = None
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, hmap["Deck_ID"]).value == args.deck and ws.cell(r, hmap["Slide_ID"]).value == sid:
                row_idx = r
                break

        if row_idx is None:
            missing.append(sid)
            continue

        def set_if(col_name: str, value):
            if value is None:
                return
            if col_name not in hmap:
                return
            ws.cell(row_idx, hmap[col_name]).value = value

        set_if("Status", args.status)
        set_if("P0", args.p0)
        set_if("PDF_OK", args.pdf_ok)
        set_if("Evidence_OK", args.evidence_ok)
        set_if("Visual_0_10", args.visual)
        set_if("Last_Reviewed", reviewed)

        if args.notes and "Notes" in hmap:
            cur = ws.cell(row_idx, hmap["Notes"]).value
            cur_s = (str(cur).strip() + " | ") if (cur and str(cur).strip()) else ""
            ws.cell(row_idx, hmap["Notes"]).value = cur_s + args.notes.strip()

        if args.title_from_html and "Title" in hmap:
            # Prefer File_Path if present; else assume ./<deck>/src/slides/<sid>.html
            slide_path = None
            if "File_Path" in hmap:
                fp = ws.cell(row_idx, hmap["File_Path"]).value
                if isinstance(fp, str) and fp.strip():
                    slide_path = Path(fp)
                    if not slide_path.is_absolute():
                        slide_path = (xlsx_path.parent / slide_path).resolve()
            if slide_path is None:
                slide_path = (xlsx_path.parent / args.deck / "src" / "slides" / f"{sid}.html").resolve()

            title = extract_title_from_html(slide_path)
            if title:
                ws.cell(row_idx, hmap["Title"]).value = title

        updated += 1

    wb.save(xlsx_path)

    print(f"Updated {updated} slide rows in {xlsx_path}")
    if missing:
        print("Slides not found in DASHBOARD.xlsx:", ", ".join(missing))

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
